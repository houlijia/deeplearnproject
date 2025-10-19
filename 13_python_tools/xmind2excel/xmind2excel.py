import json
from xmindparser import xmind_to_dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def xmind_to_excel(xmind_path, excel_path):
    """将XMind文件转换为Excel文件

    参数:
        xmind_path: XMind文件路径
        excel_path: 输出的Excel文件路径
    """
    # 1. 解析XMind文件
    try:
        xmind_data = xmind_to_dict(xmind_path)
        if not xmind_data:
            raise ValueError("XMind文件内容为空")
    except Exception as e:
        print(f"解析XMind失败: {e}")
        return False

    # 2. 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "XMind转换结果"

    # 3. 设置表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # 4. 定义表头 (可根据需要调整)
    headers = ["ID", "层级路径", "节点标题", "优先级", "备注", "预期结果"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # 5. 递归遍历节点
    row_index = 2
    root_topic = xmind_data[0]['topic']
    row_index = _parse_topic(ws, root_topic, row_index, "1", root_topic['title'])

    # 6. 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # 7. 保存Excel文件
    try:
        wb.save(excel_path)
        print(f"转换成功! 已保存到: {excel_path}")
        return True
    except Exception as e:
        print(f"保存Excel失败: {e}")
        return False


def _parse_topic(ws, topic, row_index, parent_id, parent_path):
    """递归解析XMind节点

    参数:
        ws: Excel工作表对象
        topic: 当前节点数据
        row_index: 当前行索引
        parent_id: 父节点ID
        parent_path: 父节点路径
    """
    current_id = f"{parent_id}.{row_index}"
    current_path = f"{parent_path} > {topic['title']}"

    # 提取节点属性
    title = topic.get('title', '')
    note = topic.get('note', '')
    labels = ", ".join(topic.get('labels', []))

    # 解析优先级 (根据maker标记)
    priority = "中"
    if 'makers' in topic:
        if 'priority-1' in topic['makers']:
            priority = "高"
        elif 'priority-2' in topic['makers']:
            priority = "中"
        elif 'priority-3' in topic['makers']:
            priority = "低"

    # 写入当前节点数据
    ws.cell(row=row_index, column=1, value=current_id)
    ws.cell(row=row_index, column=2, value=current_path)
    ws.cell(row=row_index, column=3, value=title)
    ws.cell(row=row_index, column=4, value=priority)
    ws.cell(row=row_index, column=5, value=note.replace('\n', ' ') if note else '')

    # 处理预期结果 (如果有特定子节点)
    expected_result = ""
    if 'topics' in topic:
        for subtopic in topic['topics']:
            if subtopic.get('title', '').lower() == "预期结果":
                expected_result = subtopic.get('note', '') or subtopic.get('title', '')
    ws.cell(row=row_index, column=6, value=expected_result)

    current_row = row_index
    row_index += 1

    # 递归处理子节点
    if 'topics' in topic:
        for subtopic in topic['topics']:
            # 跳过特定子节点类型 (如预期结果)
            if subtopic.get('title', '').lower() not in ["预期结果", "实际结果"]:
                row_index = _parse_topic(ws, subtopic, row_index, current_id, current_path)

    return row_index


# 使用示例
if __name__ == "__main__":
    xmind_to_excel("wtest111.xmind", "测试用例.xlsx")
